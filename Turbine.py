import matplotlib.pyplot as plt, pandas as pd, tkinter as tk, numpy, math, random, copy
from tkinter import filedialog, Button
from openpyxl import Workbook
from openpyxl import load_workbook
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

def initial_population(population_size):
  population = []

  for i in range(population_size):
    chromosome = numpy.arange(blades_num)
    numpy.random.shuffle(chromosome)
    population.append(chromosome)

  return population
def compute_fitness(individual):

  Mx = My = j = 0

  for i in numpy.array(individual).astype('int'):
    Mx += blades_mass[i]*blades_arm[i]*math.cos((2*math.pi*j)/len(individual))
    My += blades_mass[i]*blades_arm[i]*math.sin((2*math.pi*j)/len(individual))
    j += 1

  fitness = 1/(math.sqrt((Mx+Me*math.cos(Me_angle))**2+(My+Me*math.sin(Me_angle))**2))

  return fitness
def selection(population):
  population_fitness = sum([compute_fitness(population[i]) for i in range(len(population))])
  chromosome_probability = [(compute_fitness(population[i]))/population_fitness for i in range(len(population))]
  j = 0
  parents = []

  while j < 2:
    parents.append(population[numpy.random.choice(numpy.arange(len(population)), p=chromosome_probability)])
    j += 1
    if j == 2:
      if numpy.array_equal(parents[0], parents[1]):
        parents.remove(parents[1])
        j = 1

  return parents
def crossover(parents):
  adjacent_blades_p1 = []
  adjacent_blades_p2 = []
  union = []
  intersection = []

  for i, j in enumerate(parents[0]):
    if i == 0:
      adjacent_blades_p1.append((parents[0][len(parents[0]) - 1], parents[0][i + 1]))
      adjacent_blades_p2.append((parents[1][len(parents[0]) - 1], parents[1][i + 1]))
    elif i == len(parents[0]) - 1:
      adjacent_blades_p1.append((parents[0][i - 1], parents[0][0]))
      adjacent_blades_p2.append((parents[1][i - 1], parents[0][0]))
    else:
      adjacent_blades_p1.append((parents[0][i - 1], parents[0][i + 1]))
      adjacent_blades_p2.append((parents[1][i - 1], parents[1][i + 1]))

  for i in range(blades_num):
    union.append(set(adjacent_blades_p1[i]).union(adjacent_blades_p2[i]))
    intersection.append(set(adjacent_blades_p1[i]).intersection(adjacent_blades_p2[i]))

  child = []
  aux_union = copy.deepcopy(union)
  gene = parents[0][0]
  child.append(gene)

  while len(child) < blades_num:
    [aux_union[i].discard(gene) for i in range(blades_num)]
    
    if aux_union[gene] != set():
      aux_gene = []
      aux_len = []

      for i in aux_union[gene]:
        aux_gene.append(i)
        aux_len.append(len(union[i]))

      aux_index = numpy.flatnonzero(numpy.array(aux_len) == numpy.array(aux_len).min())
      common_neighbor = []

      for i in aux_index:
        if bool(aux_gene[i] in intersection[gene]):
          common_neighbor.append(aux_gene[i])

      if len(common_neighbor) == 1:
        gene = common_neighbor[0]
      else:
        gene = random.choice(list(set([aux_gene[i] for i in aux_index]) - set(child)))

    else:
      gene = random.choice(list(set([i for i in range(0, blades_num)]) - set(child)))

    child.append(gene)

  return child
def mutation(parents):
  mutations = []

  for k in range(2):
    i = random.randint(0, blades_num - 1)
    j = random.randint(0, blades_num - 1)
    parents[k][i], parents[k][j] = parents[k][j], parents[k][i]
    mutations.append(parents[k])

  return mutations
def Main():
  global blades_mass, blades_arm, blades_num, Me, Me_angle, wb, sheet, bar

  file_path = filedialog.askopenfilename()
  df = pd.read_excel (file_path)
  wb = load_workbook(file_path)
  sheet = wb.active

  blades_mass = df['blades_mass'].to_numpy()
  blades_arm = df['blades_arm'].to_numpy()

  blades_num_df = df.iat[1, 5]
  blades_num = numpy.array(blades_num_df).astype('int')

  Me = df.iat[11, 5]
  Me_angle = df.iat[12, 5]
  iter_aux = 100

  population = initial_population(30)
  iter = 0
  xi = 1000
  solution = []
  x = []
  it = []
  while iter < iter_aux:
    iter += 1
    new_gen = []
    while len(new_gen) < 30:
      parents = selection(population)
      child = crossover(parents)
      mutations = mutation(parents)
      new_gen.append(child)
      new_gen.append(mutations[0])
      new_gen.append(mutations[1])
    population = copy.deepcopy(new_gen)
    fit = [compute_fitness(new_gen[i]) for i in range(len(new_gen))]
    fitness = max(fit)
    index_max = max(range(len(fit)), key=fit.__getitem__)
    if 1/fitness < xi:
      xi = 1/fitness
      solution = new_gen[index_max]

    x.append(xi)
    it.append(iter)

  for i in range(blades_num):
    sheet.cell(row=i+2, column=9).value = solution[i] + 1
  sheet.cell(row=1, column=17).value = xi
  wb.save('Solution.xlsx')

  figure = Figure(figsize=(4,3), dpi=100)
  subplot = figure.add_subplot(111)
  subplot.bar(it,x,color = 'lightsteelblue')
  bar = FigureCanvasTkAgg(figure, root)
  bar.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH, expand=0)
  subplot.plot(it, x, color='green', linestyle='dashed', linewidth = 3, marker='o', markerfacecolor='blue', markersize=1)

root = tk.Tk()

canvas = tk.Canvas(root, width = 800, height = 300)
canvas.pack()

label = tk.Label(root, text='Turbine Blades Balance Model')
label.config(font=('Arial', 20))
canvas.create_window(400, 50, window=label)

browseButton_Excel = tk.Button (text='Load File...', command=Main, bg='green', fg='white', font=('helvetica', 12, 'bold'))
canvas.create_window(400, 180, window=browseButton_Excel)

exitButton = tk.Button (root, text='Exit!', command=root.destroy, bg='green', font=('helvetica', 11, 'bold'))
canvas.create_window(400, 260, window=exitButton)

root.mainloop()